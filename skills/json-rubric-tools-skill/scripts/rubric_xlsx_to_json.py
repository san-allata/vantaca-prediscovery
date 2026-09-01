#!/usr/bin/env python3
"""
Export the Assessment sheet of an Associa Branch Readiness rubric workbook to JSON.

Usage:
    python3 rubric_xlsx_to_json.py <workbook.xlsx> [-o output.json]

Design notes
------------
* Columns are located by HEADER SUBSTRING on row 4, never by fixed letter, because
  the real headers carry em-dash suffixes and future revisions may shift columns.
* Cap # is text in some rows and float in others across this rubric family, so it is
  always normalised to a string.
* Columns M/N/O (AC#/FB#/PC#) are live formulas keyed on the Classification column.
  They are deliberately NOT exported -- they are derived, not source data.
* Rows whose Discovery Question begins with "TBD" are real rows with no authored
  question; they are exported with is_placeholder_question=true so downstream code
  can skip them instead of writing "Not Found" against a non-question.
"""

import argparse
import datetime as dt
import json
import re
import sys

import openpyxl

SHEET = "Assessment"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# field name -> substring that must appear in the row-4 header
HEADER_PATTERNS = {
    "dom_number": "Dom #",
    "domain": "Domain",
    "cap_number": "Cap #",
    "capability": "Capability",          # matched after TownSq Capability is excluded
    "dimension": "Dimension",
    "priority": "Priority",              # matched after FB Priority is excluded
    "discovery_question": "Discovery Question",
    "branch_answer": "Branch Answer",
    "townsq_capability": "TownSq Capability",
    "classification": "Classification",
    "fg_priority": "FB Priority",
    "assessor_notes": "Assessor Notes",
    "qid": "QID",
    "uid": "UID",
}


def clean(value):
    """Normalise a cell value: trim whitespace, collapse empties to None."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text or None
    return value


def resolve_columns(ws):
    """Map field names to 1-based column indexes using row-4 header substrings."""
    headers = {
        col: str(ws.cell(HEADER_ROW, col).value or "").replace("\n", " ").strip()
        for col in range(1, ws.max_column + 1)
    }

    resolved = {}
    # Longest / most specific patterns first so "TownSq Capability" wins over
    # "Capability" and "FB Priority" wins over "Priority".
    order = sorted(HEADER_PATTERNS.items(), key=lambda kv: -len(kv[1]))
    taken = set()
    for field, pattern in order:
        match = next(
            (c for c, h in headers.items() if pattern in h and c not in taken), None
        )
        if match is None:
            raise SystemExit(f"Header not found for field {field!r} (pattern {pattern!r})")
        resolved[field] = match
        taken.add(match)
    return resolved, headers


def export(path, out_path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"No {SHEET!r} sheet in {path}")
    ws = wb[SHEET]

    cols, headers = resolve_columns(ws)
    qcol = cols["discovery_question"]

    rows = [
        r
        for r in range(FIRST_DATA_ROW, ws.max_row + 1)
        if clean(ws.cell(r, qcol).value) is not None
    ]
    if rows and len(rows) != max(rows) - min(rows) + 1:
        print("WARNING: question rows are not contiguous", file=sys.stderr)

    questions = []
    for r in rows:
        get = lambda field: clean(ws.cell(r, cols[field]).value)  # noqa: E731

        dimension_raw = get("dimension") or ""
        is_sub = bool(re.match(r"^\s*↳", str(ws.cell(r, cols["dimension"]).value or "")))
        dimension = re.sub(r"^\s*↳\s*", "", dimension_raw).strip() or None

        question_text = get("discovery_question")
        cap_number = get("cap_number")

        questions.append(
            {
                "uid": get("uid"),
                "qid": get("qid"),
                "source_row": r,
                "domain_number": int(get("dom_number")) if get("dom_number") is not None else None,
                "domain": get("domain"),
                "capability_number": str(cap_number) if cap_number is not None else None,
                "capability": get("capability"),
                "dimension": dimension,
                "is_sub_dimension": is_sub,
                "priority": get("priority"),
                "discovery_question": question_text,
                "is_placeholder_question": bool(
                    question_text and str(question_text).upper().startswith("TBD")
                ),
                "branch_answer": get("branch_answer"),
                "townsq_capability": get("townsq_capability"),
                "classification": get("classification"),
                "fg_priority": get("fg_priority"),
                "assessor_notes": get("assessor_notes"),
            }
        )

    def count(field):
        return sum(1 for q in questions if q[field] not in (None, ""))

    domains = {}
    for q in questions:
        key = f"{q['domain_number']}. {q['domain']}"
        domains[key] = domains.get(key, 0) + 1

    payload = {
        "schema_version": "1.0",
        "source": {
            "workbook": path.split("/")[-1],
            "sheet": SHEET,
            "header_row": HEADER_ROW,
            "first_data_row": FIRST_DATA_ROW,
            "exported_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
            "column_map": {f: openpyxl.utils.get_column_letter(c) for f, c in cols.items()},
            "excluded_columns": {
                "AC#/FB#/PC#": "live formulas derived from classification -- regenerated in Excel, not source data"
            },
        },
        "key": {
            "primary": "uid",
            "note": "uid is the permanent key and survives renumbering; qid (Dom.Cap.Seq) and source_row do not.",
        },
        "enums": {
            "classification": ["PC", "AC", "FB", "NA", None],
            "priority": ["P0", "P1", "P2", "P3"],
        },
        "editable_fields": [
            "branch_answer",
            "townsq_capability",
            "classification",
            "fg_priority",
            "assessor_notes",
        ],
        "read_only_fields": [
            "uid",
            "qid",
            "source_row",
            "domain_number",
            "domain",
            "capability_number",
            "capability",
            "dimension",
            "is_sub_dimension",
            "priority",
            "discovery_question",
        ],
        "counts": {
            "questions": len(questions),
            "placeholder_questions": sum(1 for q in questions if q["is_placeholder_question"]),
            "sub_dimension_rows": sum(1 for q in questions if q["is_sub_dimension"]),
            "branch_answer_filled": count("branch_answer"),
            "townsq_capability_filled": count("townsq_capability"),
            "classification_filled": count("classification"),
            "by_domain": domains,
        },
        "questions": questions,
    }

    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
        fh.write("\n")

    return payload


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("-o", "--output", default="rubric.json")
    args = ap.parse_args()
    payload = export(args.workbook, args.output)
    c = payload["counts"]
    print(f"Wrote {args.output}: {c['questions']} questions, "
          f"{c['placeholder_questions']} placeholder, "
          f"H filled {c['branch_answer_filled']}, "
          f"I filled {c['townsq_capability_filled']}, "
          f"J filled {c['classification_filled']}")


if __name__ == "__main__":
    main()
