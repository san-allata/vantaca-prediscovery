#!/usr/bin/env python3
"""
Write rubric JSON back into the original Associa Branch Readiness workbook.

Usage:
    python3 rubric_json_to_xlsx.py rubric.json --template original.xlsx -o updated.xlsx
    python3 rubric_json_to_xlsx.py rubric.json --template original.xlsx --overwrite
    python3 rubric_json_to_xlsx.py rubric.json --template original.xlsx --no-recalc

Why a template is required
--------------------------
The workbook is not reconstructed from scratch. The original file is opened with
formulas intact and only the five editable cells per row are overwritten, so all
formatting, merged banners, data validation, autofilter, column widths and the live
formula columns survive byte-for-byte everywhere this script does not touch.

Safety rails
------------
* Rows are matched by UID, never by source_row. A renumbered or re-sorted workbook
  would silently shift row indexes; UID is the permanent key.
* Before writing, the JSON discovery_question is compared against the target row's
  column G. Any mismatch aborts the run -- row misalignment is a critical defect.
* Columns A-G (read-only source) and M/N/O (live formulas) are never written.
* Placeholder rows ("TBD -- confirm question...") are skipped by default so nothing
  is written against a non-question.
* --only-empty (the default) never clobbers an existing answer. Use --overwrite to
  force, which still refuses to blank a populated cell with a null.
* After saving, a LibreOffice headless pass recalculates so the M/N/O columns and the
  formula-driven Change Plan / Onboarding List / Product Backlog / Summary sheets carry
  cached values. Without it those cells read back as None to pandas and most previewers.
* The recalc round-trip has been known to hide rows on this workbook family, so the
  saved XML is checked for hidden="1" on every sheet before the file is declared good.

Known limitation: an openpyxl round-trip drops threaded comments. If the original
carries reviewer comments, re-add them from the original after this pass.
"""

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl

SHEET = "Assessment"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# JSON field -> header substring. These are the ONLY cells this script writes.
EDITABLE = {
    "branch_answer": "Branch Answer",
    "townsq_capability": "TownSq Capability",
    "classification": "Classification",
    "fg_priority": "FB Priority",
    "assessor_notes": "Assessor Notes",
}
QUESTION_HEADER = "Discovery Question"
UID_HEADER = "UID"

VALID_CLASSIFICATION = {"PC", "AC", "FB", "NA"}
VALID_PRIORITY = {"P0", "P1", "P2", "P3"}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text or None
    return value


def norm(text):
    """Loose comparison form for the row-alignment guard."""
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


def resolve_columns(ws):
    headers = {
        col: str(ws.cell(HEADER_ROW, col).value or "").replace("\n", " ").strip()
        for col in range(1, ws.max_column + 1)
    }
    wanted = dict(EDITABLE)
    wanted["_question"] = QUESTION_HEADER
    wanted["_uid"] = UID_HEADER

    resolved, taken = {}, set()
    # Most specific patterns first: "TownSq Capability" before "Capability",
    # "FB Priority" before "Priority".
    for field, pattern in sorted(wanted.items(), key=lambda kv: -len(kv[1])):
        match = next((c for c, h in headers.items() if pattern in h and c not in taken), None)
        if match is None:
            raise SystemExit(f"Header not found for {field!r} (pattern {pattern!r})")
        resolved[field] = match
        taken.add(match)
    return resolved


def validate(questions):
    """Fail before touching the workbook, not halfway through it."""
    problems = []
    seen = set()
    for i, q in enumerate(questions):
        uid = q.get("uid")
        if not uid:
            problems.append(f"question[{i}] has no uid")
            continue
        if uid in seen:
            problems.append(f"duplicate uid {uid}")
        seen.add(uid)

        cls = clean(q.get("classification"))
        if cls is not None and str(cls).upper() not in VALID_CLASSIFICATION:
            problems.append(f"{uid}: classification {cls!r} not in {sorted(VALID_CLASSIFICATION)}")

        fgp = clean(q.get("fg_priority"))
        if fgp is not None and str(fgp).upper() not in VALID_PRIORITY:
            problems.append(f"{uid}: fg_priority {fgp!r} not in {sorted(VALID_PRIORITY)}")

        if fgp is not None and (cls is None or str(cls).upper() != "FB"):
            problems.append(f"{uid}: fg_priority set but classification is not FB")
    return problems


def recalc(path):
    """Force LibreOffice to compute and cache every formula result."""
    path = Path(path).resolve()
    with tempfile.TemporaryDirectory() as tmp:
        proc = subprocess.run(
            ["soffice", "--headless", "--norestore", "--convert-to", "xlsx",
             "--outdir", tmp, str(path)],
            capture_output=True, text=True, timeout=600,
        )
        produced = Path(tmp) / path.name
        if not produced.exists():
            raise SystemExit(f"recalc failed: {proc.stdout}\n{proc.stderr}")
        shutil.move(str(produced), str(path))


def assert_no_hidden_rows(path):
    """The recalc round-trip has silently hidden rows on this workbook family."""
    offenders = []
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                xml = z.read(name).decode("utf-8", "replace")
                hidden = len(re.findall(r'<row[^>]*hidden="1"', xml))
                if hidden:
                    offenders.append(f"{name}: {hidden} hidden rows")
    return offenders


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("json_file")
    ap.add_argument("--template", required=True,
                    help="original .xlsx to update in place (formatting/formulas preserved)")
    ap.add_argument("-o", "--output", default=None)
    ap.add_argument("--overwrite", action="store_true",
                    help="replace existing cell values (default: fill only empty cells)")
    ap.add_argument("--include-placeholders", action="store_true",
                    help="also write rows whose question is a 'TBD' placeholder")
    ap.add_argument("--no-recalc", action="store_true",
                    help="skip the LibreOffice recalc pass (M/N/O will have no cached values)")
    ap.add_argument("--dry-run", action="store_true", help="report planned writes, save nothing")
    args = ap.parse_args()

    payload = json.load(open(args.json_file, encoding="utf-8"))
    questions = payload["questions"] if isinstance(payload, dict) else payload

    problems = validate(questions)
    if problems:
        print("Refusing to write -- JSON validation failed:", file=sys.stderr)
        for p in problems[:40]:
            print("  -", p, file=sys.stderr)
        if len(problems) > 40:
            print(f"  ... and {len(problems) - 40} more", file=sys.stderr)
        raise SystemExit(1)

    out = Path(args.output) if args.output else Path(args.template)
    if not args.dry_run and out != Path(args.template):
        shutil.copyfile(args.template, out)
    target = args.template if args.dry_run else str(out)

    # No data_only -- keep the formulas in M/N/O and on the derived sheets alive.
    wb = openpyxl.load_workbook(target)
    ws = wb[SHEET]
    cols = resolve_columns(ws)

    # UID -> row index, from the workbook itself.
    uid_col, q_col = cols["_uid"], cols["_question"]
    row_by_uid = {}
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        uid = clean(ws.cell(r, uid_col).value)
        if uid is not None:
            row_by_uid[str(uid)] = r

    writes = 0
    skipped_existing = 0
    skipped_placeholder = 0
    missing = []
    misaligned = []
    per_field = {f: 0 for f in EDITABLE}

    for q in questions:
        uid = str(q["uid"])
        row = row_by_uid.get(uid)
        if row is None:
            missing.append(uid)
            continue

        # Row-alignment guard: the question in the JSON must be the question in that row.
        sheet_q = norm(ws.cell(row, q_col).value)
        json_q = norm(q.get("discovery_question"))
        if json_q and sheet_q and json_q != sheet_q:
            misaligned.append(f"{uid} (row {row})")
            continue

        if q.get("is_placeholder_question") and not args.include_placeholders:
            skipped_placeholder += 1
            continue

        for field, _ in EDITABLE.items():
            value = clean(q.get(field))
            if value is None:
                continue
            if field in ("classification", "fg_priority"):
                value = str(value).upper()
            cell = ws.cell(row, cols[field])
            existing = clean(cell.value)
            if existing is not None and not args.overwrite:
                skipped_existing += 1
                continue
            if existing == value:
                continue
            cell.value = value
            writes += 1
            per_field[field] += 1

    if misaligned:
        print("Refusing to write -- UID/question mismatch (workbook and JSON disagree):",
              file=sys.stderr)
        for m in misaligned[:20]:
            print("  -", m, file=sys.stderr)
        raise SystemExit(1)

    print(f"Matched {len(questions) - len(missing)}/{len(questions)} questions by UID")
    if missing:
        print(f"  {len(missing)} uid(s) not in workbook: {missing[:10]}")
    print(f"Planned writes: {writes}  " + "  ".join(f"{k}={v}" for k, v in per_field.items() if v))
    print(f"Skipped: {skipped_existing} already-populated cell(s), "
          f"{skipped_placeholder} placeholder row(s)")

    if args.dry_run:
        print("Dry run -- nothing saved.")
        return

    wb.save(target)

    if not args.no_recalc:
        recalc(target)

    hidden = assert_no_hidden_rows(target)
    if hidden:
        print("WARNING: hidden rows detected after recalc -- unhide before sharing:",
              file=sys.stderr)
        for h in hidden:
            print("  -", h, file=sys.stderr)
    else:
        print("Verified: no hidden rows on any sheet.")

    print(f"Saved {target}")
    print("Note: threaded comments do not survive an openpyxl round-trip; "
          "re-add from the original if the workbook carried reviewer comments.")


if __name__ == "__main__":
    main()
