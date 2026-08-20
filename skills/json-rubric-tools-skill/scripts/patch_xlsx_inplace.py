#!/usr/bin/env python3
"""
patch_xlsx_inplace.py

Patch ONLY specific answer cells in an EXISTING Excel workbook, preserving
every other sheet, cell, formula, and formatting exactly as-is.

This is the REQUIRED final-delivery step for rubric workflows that must
return "the original file with only answer cells updated" -- never a
freshly-rebuilt single-sheet workbook. Unlike rubric_json_to_xlsx.py (which
constructs a brand-new workbook from JSON and therefore drops any other
sheets), this script opens the client's real file with openpyxl and writes
directly into the loaded workbook object, so nothing else in the file is
touched.

Usage:
    python patch_xlsx_inplace.py \
        --input original.xlsx \
        --updates updates.json \
        --output completed.xlsx \
        [--sheet Assessment] \
        [--question-col "Discovery Question"] \
        [--answer-col "Branch Answer"] \
        [--header-row 4] \
        [--data-start-row 5]

updates.json format (array of objects). Two supported shapes, may be mixed:

1) Row-addressed (use when you already know the exact worksheet row):
[
  {"row": 5, "answer": "We reconcile 12 accounts monthly.\n\nSource: Ops Manual Section 4.1"},
  {"row": 6, "answer": "Not Found -- no lockbox details in materials.\n\nSource: Ops Manual Section 3"}
]

2) Question-addressed (tool resolves the row as data_start_row + question_index):
[
  {"question_index": 0, "answer": "..."},
  {"question_index": 2, "answer": "..."}
]

IMPORTANT: This script performs NO answer formatting or reformatting.
Callers (the rubric-answer-extractor-integrated skill / json-question-answer-patcher)
must pre-format every `answer` string exactly per references/answer-format.md
before calling this script, e.g.:

    f"{answer_corpus}\n\nSource: {source}"

    # partial match: append gaps
    f"{answer_corpus}\n\nSource: {source}\n\nGAPS: {gaps}"

The script prints a JSON summary to stdout: output path, sheet, resolved
answer column, rows updated, and any updates that could not be resolved.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl is required but not installed"}))
    sys.exit(1)


def find_header_columns(ws, header_row, question_col_substr, answer_col_substr):
    q_col = None
    a_col = None
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        val = str(cell.value)
        if question_col_substr.lower() in val.lower():
            q_col = cell.column
        if answer_col_substr.lower() in val.lower():
            a_col = cell.column
    return q_col, a_col


def main():
    parser = argparse.ArgumentParser(
        description="Patch answer cells into an existing xlsx in place, preserving all other sheets/formatting."
    )
    parser.add_argument("--input", required=True, help="Path to the original .xlsx file")
    parser.add_argument("--updates", required=True, help="Path to JSON file with the list of updates")
    parser.add_argument("--output", default=None, help="Output path (defaults to <input>_completed.xlsx)")
    parser.add_argument("--sheet", default="Assessment", help="Target sheet name (default: Assessment)")
    parser.add_argument("--question-col", default="Discovery Question", help="Substring to find the question header")
    parser.add_argument("--answer-col", default="Branch Answer", help="Substring to find the answer header")
    parser.add_argument("--header-row", type=int, default=4, help="Row number containing headers (default: 4)")
    parser.add_argument("--data-start-row", type=int, default=5, help="First data row (default: 5)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(json.dumps({"error": f"Input file not found: {input_path}"}))
        sys.exit(1)

    with open(args.updates, "r", encoding="utf-8") as f:
        updates = json.load(f)

    output_path = Path(args.output) if args.output else input_path.with_name(
        input_path.stem + "_completed" + input_path.suffix
    )

    # Loading (not rebuilding) the workbook is what preserves every other
    # sheet, formula, chart, and formatting untouched.
    wb = load_workbook(filename=str(input_path))

    if args.sheet not in wb.sheetnames:
        print(json.dumps({
            "error": f"Sheet '{args.sheet}' not found. Available sheets: {wb.sheetnames}"
        }))
        sys.exit(1)

    ws = wb[args.sheet]
    q_col, a_col = find_header_columns(ws, args.header_row, args.question_col, args.answer_col)
    if a_col is None:
        print(json.dumps({
            "error": f"Could not locate answer column header containing '{args.answer_col}' in row {args.header_row}"
        }))
        sys.exit(1)

    updated_rows = []
    skipped = []

    for upd in updates:
        row = upd.get("row")
        answer = upd.get("answer")
        if answer is None:
            skipped.append({"update": upd, "reason": "missing 'answer'"})
            continue

        if row is None and "question_index" in upd:
            target_row = args.data_start_row + int(upd["question_index"])
        elif row is not None:
            target_row = int(row)
        else:
            skipped.append({"update": upd, "reason": "no row or question_index resolvable"})
            continue

        if target_row < args.data_start_row or target_row > ws.max_row:
            skipped.append({"update": upd, "reason": f"row {target_row} out of data range (max_row={ws.max_row})"})
            continue

        ws.cell(row=target_row, column=a_col, value=answer)
        updated_rows.append(target_row)

    wb.save(str(output_path))

    print(json.dumps({
        "output": str(output_path),
        "sheet": args.sheet,
        "answer_column": a_col,
        "rows_updated": updated_rows,
        "rows_updated_count": len(updated_rows),
        "skipped": skipped,
        "sheets_preserved": wb.sheetnames,
    }, indent=2))


if __name__ == "__main__":
    main()
