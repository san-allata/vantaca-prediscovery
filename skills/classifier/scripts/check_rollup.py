#!/usr/bin/env python3
"""
check_rollup.py — verify the workbook actually reflects the classifications.

Read-only. Repairs nothing. Run it after converting classified JSON back to
.xlsx, because a classification nobody can see at the domain level has not been
delivered.

Six checks, each of which has caught a real defect on this workbook family:

  1. summary_column   Summary counts must reference the Classification column,
                      not the adjacent capability column. Pointing at the wrong
                      column makes every count read zero while looking fine.
  2. summary_range    Summary ranges must cover the full data extent, not a
                      stale lower bound from a shorter rubric revision.
  3. domain_roster    The Summary's domain list must match the domain numbers
                      actually present on the Assessment sheet.
  4. reconcile        Summary totals must equal an independent recount of the
                      Classification column.
  5. helper_columns   The AC#/FB#/PC# helper formulas must be anchored to their
                      OWN row. They feed the PC / AC / FB worklist sheets by
                      MATCH, so a mis-anchored or missing formula silently drops
                      real rows from a worklist and inserts unrelated ones.
  6. artefacts        Hidden rows (a recalc round-trip has hidden rows here
                      before) and a comments part (threaded comments do not
                      survive an openpyxl round-trip).

Exit 1 if any check fails.

Usage:
    python scripts/check_rollup.py workbook.xlsx [--verbose]
"""
import argparse, re, sys, zipfile
from collections import Counter

import openpyxl

SHEET = "Assessment"
HEADER_ROW = 4
FIRST_DATA_ROW = 5
CODES = ["PC", "AC", "FB", "NA"]
WORKLISTS = {"PC": "Change Plan", "AC": "Onboarding List", "FB": "Product Backlog"}


def col_of(ws, pattern, taken=()):
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(HEADER_ROW, c).value or "").replace("\n", " ").strip()
        if pattern in h and c not in taken:
            return c
    return None


def letter(c):
    return openpyxl.utils.get_column_letter(c)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    wf = openpyxl.load_workbook(args.workbook)              # formulas
    wv = openpyxl.load_workbook(args.workbook, data_only=True)  # cached values
    if SHEET not in wf.sheetnames:
        sys.exit(f"no {SHEET!r} sheet")
    af, av = wf[SHEET], wv[SHEET]

    # resolve the columns that matter, most specific pattern first
    taken = set()
    cls_col = col_of(af, "Classification")
    taken.add(cls_col)
    q_col = col_of(af, "Discovery Question", taken)
    taken.add(q_col)
    dom_col = col_of(af, "Dom #", taken)
    taken.add(dom_col)
    helper = {}
    for code in ("AC", "FB", "PC"):
        c = col_of(af, f"{code}#", taken)
        if c:
            helper[code] = c
            taken.add(c)

    last = max(r for r in range(FIRST_DATA_ROW, af.max_row + 1) if af.cell(r, q_col).value)
    fails, notes = [], []

    print(f"WORKBOOK {args.workbook}")
    print(f"  sheet {SHEET}: data rows {FIRST_DATA_ROW}-{last} ({last - FIRST_DATA_ROW + 1})")
    print(f"  Classification = column {letter(cls_col)}   "
          f"helpers = " + ", ".join(f"{k}:{letter(v)}" for k, v in helper.items()))

    # --- independent recount ---
    recount = Counter()
    for r in range(FIRST_DATA_ROW, last + 1):
        v = av.cell(r, cls_col).value
        if v is not None and str(v).strip():
            recount[str(v).strip().upper()] += 1
    off_enum = {k: n for k, n in recount.items() if k not in CODES}
    print("\n[recount] column " + letter(cls_col) + ": " +
          "  ".join(f"{k}={recount.get(k, 0)}" for k in CODES) +
          f"   total classified {sum(recount.values())}")
    if off_enum:
        fails.append(f"non-code values in the classification column: {off_enum} "
                     f"(these look classified in the grid and count as unclassified everywhere else)")

    # --- codes with no branch answer are unsupported ---
    ba = col_of(af, "Branch Answer")
    unsupported = [r for r in range(FIRST_DATA_ROW, last + 1)
                   if av.cell(r, cls_col).value and not av.cell(r, ba).value]
    if unsupported:
        notes.append(f"{len(unsupported)} row(s) carry a code with no branch answer "
                     f"(rows {unsupported[:8]}) — nothing was compared against")

    # --- Summary sheet ---
    sname = next((n for n in wf.sheetnames if n.lower().startswith("summary")), None)
    if not sname:
        fails.append("no Summary sheet found")
    else:
        sf, sv = wf[sname], wv[sname]
        refs = []
        for row in sf.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and SHEET in v:
                    refs.append((cell.coordinate, v))
        code_refs = [(c, v) for c, v in refs if any(f'"{k}"' in v for k in CODES)]
        # Only the range compared DIRECTLY to a code literal matters. These formulas
        # also carry a domain-number range (e.g. $A$5:$A$730=$A7) which is correct and
        # must not be flagged.
        pat = re.compile(
            rf"{re.escape(SHEET)}!\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)\s*=\s*\"({'|'.join(CODES)})\"")
        wrong_col, code_ranges = [], set()
        for coord, v in code_refs:
            for a, r1, r2, _code in pat.findall(v):
                if openpyxl.utils.column_index_from_string(a) != cls_col:
                    wrong_col.append(a)
                else:
                    code_ranges.add((a, int(r1), int(r2)))
        if not code_refs:
            fails.append("Summary has no formula counting classification codes")
        elif not code_ranges and not wrong_col:
            notes.append("could not parse the Summary's code ranges; verify by hand")
        print(f"\n[summary] sheet {sname!r}: {len(code_refs)} code-counting formula(s)")
        if code_ranges:
            print("  code ranges: " + ", ".join(f"{a}{r1}:{a}{r2}" for a, r1, r2 in sorted(code_ranges)))
        if wrong_col:
            fails.append(f"Summary code counts reference column(s) {sorted(set(wrong_col))} "
                         f"but Classification is column {letter(cls_col)} — every count will read zero")
        short = [(a, r1, r2) for a, r1, r2 in code_ranges if r2 < last]
        if short:
            fails.append(f"Summary code ranges stop at row {min(r2 for _, _, r2 in short)} "
                         f"but data runs to {last} — rows beyond are invisible to the roll-up")
        elif code_ranges:
            print(f"  range covers the full extent (data ends {last})")

        # roster
        present = {int(av.cell(r, dom_col).value) for r in range(FIRST_DATA_ROW, last + 1)
                   if av.cell(r, dom_col).value is not None}
        listed = set()
        for r in range(1, sv.max_row + 1):
            v = sv.cell(r, 1).value
            if isinstance(v, (int, float)) and float(v).is_integer():
                listed.add(int(v))
        missing, extra = sorted(present - listed), sorted(listed - present)
        if missing or extra:
            fails.append(f"Summary domain roster mismatch — missing {missing}, "
                         f"lists absent {extra}")
        else:
            print(f"  domain roster matches the {len(present)} domains present")

        # reconcile totals
        tot = {}
        for r in range(1, sv.max_row + 1):
            if str(sv.cell(r, 2).value or "").strip().upper() == "TOTAL":
                for c in range(1, sv.max_column + 1):
                    hdr = str(sv.cell(6, c).value or "").strip().upper()
                    if hdr in CODES:
                        tot[hdr] = sv.cell(r, c).value
                break
        if tot:
            bad = {k: (tot.get(k), recount.get(k, 0)) for k in CODES
                   if (tot.get(k) or 0) != recount.get(k, 0)}
            if bad:
                fails.append(f"Summary totals do not reconcile with a recount "
                             f"(summary, recount): {bad}")
            else:
                print("  totals reconcile with the recount: " +
                      "  ".join(f"{k}={tot.get(k)}" for k in CODES))

    # --- helper columns ---
    print("\n[helpers] AC#/FB#/PC# formulas feed the worklist sheets by MATCH")
    helper_broken = False
    for code, c in helper.items():
        ok = mis = miss = 0
        examples = []
        for r in range(FIRST_DATA_ROW, last + 1):
            v = af.cell(r, c).value
            if not (isinstance(v, str) and v.startswith("=")):
                miss += 1
                continue
            m = re.search(r"\$[A-Z]+(\d+)\s*=", v)
            if m and int(m.group(1)) == r:
                ok += 1
            else:
                mis += 1
                if len(examples) < 3 and m:
                    examples.append(f"row {r}->{m.group(1)} ({int(m.group(1)) - r:+d})")
        line = (f"  {code}# ({letter(c)}): anchored {ok}  mis-anchored {mis}  missing {miss}")
        print(line + ("   e.g. " + "; ".join(examples) if examples else ""))
        if mis or miss:
            helper_broken = True
    if helper_broken:
        fails.append(f"helper columns are damaged — the "
                     f"{', '.join(WORKLISTS[k] for k in helper if k in WORKLISTS)} sheets "
                     f"will drop real rows and show unrelated ones. The Summary is unaffected "
                     f"because it reads the classification column directly.")

    # --- worklist sanity: do the rows they name actually carry that code? ---
    for code, prefix in WORKLISTS.items():
        sn = next((n for n in wv.sheetnames if n.startswith(prefix)), None)
        if not sn:
            continue
        s = wv[sn]
        listed = [int(s.cell(r, 1).value) for r in range(1, s.max_row + 1)
                  if isinstance(s.cell(r, 1).value, (int, float))
                  and FIRST_DATA_ROW <= s.cell(r, 1).value <= last]
        wrong = [r for r in listed
                 if str(av.cell(r, cls_col).value or "").strip().upper() != code]
        expected = recount.get(code, 0)
        status = "OK" if not wrong and len(listed) == expected else "MISMATCH"
        print(f"\n[worklist] {sn}: lists {len(listed)} row(s), expected {expected} — {status}")
        if wrong:
            print(f"  rows listed that are not {code}: {wrong[:10]}")
        if status == "MISMATCH":
            fails.append(f"{sn} does not match the {code} rows in the classification column")

    # --- artefacts ---
    hidden, comments = [], False
    with zipfile.ZipFile(args.workbook) as z:
        for n in z.namelist():
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                cnt = len(re.findall(r'<row[^>]*hidden="1"', z.read(n).decode("utf-8", "replace")))
                if cnt:
                    hidden.append(f"{n}: {cnt}")
            if "comment" in n:
                comments = True
    print("\n[artefacts] hidden rows: " + (", ".join(hidden) if hidden else "none"))
    if hidden:
        fails.append(f"hidden rows present: {hidden} — the file will look filtered")
    if comments:
        notes.append("workbook carries a comments part; threaded comments do not survive an "
                     "openpyxl round-trip and should be re-added from the original")

    print("\n" + "=" * 62)
    if notes:
        print(f"NOTES ({len(notes)}):")
        for n in notes:
            print("  - " + n)
    if fails:
        print(f"\nFAILURES ({len(fails)}):")
        for f in fails:
            print("  - " + f)
        print("\nThe roll-up does not faithfully reflect the classifications. Repairing this "
              "is a deliberate, separately verified pass — not a side effect of a conversion.")
    else:
        print("PASS — the roll-up faithfully reflects the classification column.")
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
