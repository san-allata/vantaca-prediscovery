
import json
import sys
from openpyxl import load_workbook

def main():
    with open("input.json", "r", encoding="utf-8") as f:
        payload = json.load(f)

    excel_file = payload["excel_file"]
    sheet_name = payload.get("sheet_name", "Assessment")
    question_col = payload.get("question_col", "G")
    answer_col = payload.get("answer_col", "H")
    answers = payload["answers"]  # dict: question_text -> answer_text
    output_file = payload.get("output_file", "output_PROCESSED.xlsx")

    print("=" * 80)
    print("EXCEL Q&A PROCESSOR - COMPLETE PRESERVATION")
    print("=" * 80)

    wb = load_workbook(excel_file, data_only=False)
    print(f"Loaded workbook: {excel_file}")
    print(f"Original sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] Sheet '{sheet_name}' not found.")
        sys.exit(1)

    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Assessment sheet dimensions: {max_row} rows x {max_col} cols")

    # Normalize keys for robust matching (strip whitespace)
    norm_answers = {k.strip(): v for k, v in answers.items()}

    matched = 0
    unmatched_questions = []
    updated_rows = []

    for row in range(1, max_row + 1):
        q_cell = ws[f"{question_col}{row}"]
        q_val = q_cell.value
        if q_val is None:
            continue
        q_text = str(q_val).strip()
        if q_text in norm_answers:
            ans_cell = ws[f"{answer_col}{row}"]
            ans_cell.value = norm_answers[q_text]
            matched += 1
            updated_rows.append(row)

    # Report which provided answers were never matched to a question (helps QA)
    all_q_texts = set()
    for row in range(1, max_row + 1):
        v = ws[f"{question_col}{row}"].value
        if v is not None:
            all_q_texts.add(str(v).strip())

    for q in norm_answers:
        if q not in all_q_texts:
            unmatched_questions.append(q)

    wb.save(output_file)

    print(f"\nMatched and updated {matched} answer cells.")
    print(f"Updated rows: {updated_rows}")
    if unmatched_questions:
        print(f"\n[WARNING] {len(unmatched_questions)} provided answers had no matching question text:")
        for q in unmatched_questions:
            print(f"  - {q[:100]}")

    # Verify preservation
    wb2 = load_workbook(output_file, data_only=False)
    print(f"\nVerification - Output sheets ({len(wb2.sheetnames)}): {wb2.sheetnames}")
    ws2 = wb2[sheet_name]
    print(f"Verification - Assessment sheet dimensions: {ws2.max_row} rows x {ws2.max_column} cols")
    print("=" * 80)
    print("DONE")

if __name__ == "__main__":
    main()
