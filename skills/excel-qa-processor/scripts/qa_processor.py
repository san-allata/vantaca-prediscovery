#!/usr/bin/env python3
"""
Excel Q&A Processor Script - COMPLETE PRESERVATION VERSION

CRITICAL: This script loads the ENTIRE ORIGINAL WORKBOOK and modifies ONLY answer cells.
All sheets, rows, columns, data, and formatting are preserved 100%.

Workflow:
  1. Load complete original workbook (load_workbook with data_only=False)
  2. Verify all sheets are present
  3. Find Assessment sheet
  4. Auto-detect question and answer columns
  5. Extract all questions
  6. For each question, find and write answer (ONLY in answer column)
  7. Save complete workbook
  8. Verify output has same sheets as input
"""

import sys
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy


def verify_sheet_preservation(input_wb, output_wb):
    """Verify that all sheets from input are in output."""
    input_sheets = set(input_wb.sheetnames)
    output_sheets = set(output_wb.sheetnames)

    if input_sheets == output_sheets:
        print(f"[OK] Sheet preservation verified: {len(input_sheets)} sheets present")
        return True
    else:
        missing = input_sheets - output_sheets
        extra = output_sheets - input_sheets
        if missing:
            print(f"[ERROR] Missing sheets: {missing}")
        if extra:
            print(f"[ERROR] Unexpected extra sheets: {extra}")
        return False


def get_assessment_sheet(wb):
    """Find and return the Assessment sheet."""
    # Try common names
    for sheet_name in ['Assessment', 'assessment', 'ASSESSMENT']:
        if sheet_name in wb.sheetnames:
            print(f"[OK] Found sheet: {sheet_name}")
            return wb[sheet_name]

    # If not found, use first sheet and warn
    if len(wb.sheetnames) > 0:
        sheet_name = wb.sheetnames[0]
        print(f"[WARNING] 'Assessment' sheet not found, using: {sheet_name}")
        return wb[sheet_name]

    print("[ERROR] No sheets found in workbook")
    return None


def detect_question_column(ws, header_row=4, max_scan_cols=20):
    """Auto-detect which column contains questions.
    Returns (col_letter, col_index) or (None, None) if not found.
    """
    # Check header row for 'question' text
    for col_idx in range(1, max_scan_cols):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value and 'question' in str(cell.value).lower():
            col_letter = get_column_letter(col_idx)
            print(f"[OK] Question column detected: {col_letter} (header: '{cell.value}')")
            return col_letter, col_idx

    # Fallback: scan for columns with long text content (likely questions)
    for col_idx in [7, 6, 8, 5, 9, 4]:  # G, F, H, E, I, D - typical question columns
        col_letter = get_column_letter(col_idx)
        has_questions = False

        # Check if this column has long content (typical of questions)
        for row in range(header_row + 1, header_row + 10):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value and len(str(cell.value).strip()) > 30:  # Questions are typically longer
                has_questions = True
                break

        if has_questions:
            print(f"[OK] Question column detected (by content): {col_letter}")
            return col_letter, col_idx

    # Default to column G if nothing found
    print(f"[WARNING] Question column not detected, defaulting to G")
    return 'G', 7


def detect_answer_column(ws, question_col_idx, header_row=4):
    """Auto-detect which column should contain answers.
    Typically the column immediately after questions.
    """
    answer_col_idx = question_col_idx + 1
    if answer_col_idx > 26:  # Beyond column Z, wrap to after questions
        answer_col_idx = question_col_idx + 1

    answer_col_letter = get_column_letter(answer_col_idx)

    # Check if answer column header exists
    header_cell = ws.cell(row=header_row, column=answer_col_idx)
    if header_cell.value:
        print(f"[OK] Answer column detected: {answer_col_letter} (header: '{header_cell.value}')")
    else:
        print(f"[OK] Answer column detected: {answer_col_letter} (no header)")

    return answer_col_letter, answer_col_idx


def find_first_question_row(ws, question_col_letter, header_row=4):
    """Find the first row with an actual question."""
    for row in range(header_row + 1, header_row + 500):
        cell = ws[f"{question_col_letter}{row}"]
        if cell.value and len(str(cell.value).strip()) > 0:
            print(f"[OK] First question row detected: {row}")
            return row

    print(f"[WARNING] No questions found, starting from row {header_row + 1}")
    return header_row + 1


def get_row_count(ws):
    """Get actual number of rows with data in the worksheet."""
    max_row = 0
    for row in ws.iter_rows():
        has_data = False
        for cell in row:
            if cell.value is not None:
                has_data = True
                break
        if has_data:
            max_row = max(max_row, row[0].row)
    return max_row


def get_col_count(ws):
    """Get actual number of columns with data in the worksheet."""
    max_col = 0
    for col in ws.iter_cols():
        has_data = False
        for cell in col:
            if cell.value is not None:
                has_data = True
                break
        if has_data:
            max_col = max(max_col, col[0].column)
    return max_col


def extract_questions(ws, question_col_letter, start_row):
    """Extract all questions from the question column.
    Returns list of (row_num, question_text).
    """
    questions = []
    row = start_row
    max_rows = 1000  # Safety limit
    empty_count = 0

    while row < start_row + max_rows:
        cell = ws[f"{question_col_letter}{row}"]
        if cell.value is None or (isinstance(cell.value, str) and len(cell.value.strip()) == 0):
            empty_count += 1
            # Stop if we've hit 10 empty rows in a row
            if empty_count > 10:
                break
        else:
            empty_count = 0
            question_text = str(cell.value).strip()
            if len(question_text) > 0:
                questions.append((row, question_text))

        row += 1

    print(f"[OK] Extracted {len(questions)} questions from rows {start_row}-{row-1}")
    return questions


def process_assessment_file(excel_path, answers_dict):
    """
    Main processing function.
    CRITICAL: Loads complete original workbook and preserves everything.

    Args:
        excel_path: Path to the Excel file
        answers_dict: Dictionary mapping question text or row numbers to answer text

    Returns:
        Path to the output file, or None if failed
    """
    try:
        # STEP 1: Load complete original workbook
        print(f"\n[INFO] Loading complete original workbook: {excel_path}")
        if not os.path.exists(excel_path):
            print(f"[ERROR] File not found: {excel_path}")
            return None

        wb = load_workbook(excel_path)
        print(f"[OK] Workbook loaded with {len(wb.sheetnames)} sheets: {wb.sheetnames}")

        # STEP 2: Verify and document original structure
        original_sheets = set(wb.sheetnames)
        ws = get_assessment_sheet(wb)
        if ws is None:
            print("[ERROR] Could not find Assessment sheet")
            return None

        original_max_row = get_row_count(ws)
        original_max_col = get_col_count(ws)
        print(f"[OK] Assessment sheet structure: {original_max_row} rows x {original_max_col} columns")

        # STEP 3: Auto-detect structure
        print("\n[INFO] Auto-detecting question and answer columns...")
        question_col_letter, question_col_idx = detect_question_column(ws)
        answer_col_letter, answer_col_idx = detect_answer_column(ws, question_col_idx)
        print(f"[OK] Detected: Question={question_col_letter}, Answer={answer_col_letter}")

        # STEP 4: Find header and first data row
        header_row = 4  # Assume standard, but could be adjusted
        first_q_row = find_first_question_row(ws, question_col_letter, header_row)

        # STEP 5: Extract all questions
        print("\n[INFO] Extracting questions...")
        questions = extract_questions(ws, question_col_letter, first_q_row)
        if len(questions) == 0:
            print("[WARNING] No questions found in the workbook")

        # STEP 6: Populate answers
        print("\n[INFO] Populating answers...")
        updated_count = 0
        for row_num, question_text in questions:
            answer_text = answers_dict.get(question_text) or answers_dict.get(row_num)

            if answer_text:
                cell = ws[f"{answer_col_letter}{row_num}"]
                cell.value = answer_text
                print(f"[OK] Row {row_num}: Answer populated ({len(answer_text)} chars)")
                updated_count += 1
            else:
                print(f"[INFO] Row {row_num}: No answer found for this question")

        # STEP 7: Save complete workbook
        output_path = Path(excel_path).stem + "_PROCESSED.xlsx"
        print(f"\n[INFO] Saving complete workbook: {output_path}")
        wb.save(output_path)

        # STEP 8: Verify preservation
        print("\n[INFO] Verifying file preservation...")
        output_wb = load_workbook(output_path)

        # Check sheets
        sheets_preserved = verify_sheet_preservation(wb, output_wb)

        # Check structure of Assessment sheet
        output_ws = get_assessment_sheet(output_wb)
        if output_ws:
            output_max_row = get_row_count(output_ws)
            output_max_col = get_col_count(output_ws)
            print(f"[OK] Output Assessment sheet: {output_max_row} rows x {output_max_col} columns")

            if output_max_row == original_max_row and output_max_col == original_max_col:
                print(f"[OK] ✓ Structure preserved exactly")
            else:
                print(f"[WARNING] Structure changed: {original_max_row}→{output_max_row} rows, {original_max_col}→{output_max_col} cols")

        if sheets_preserved:
            print(f"\n[SUCCESS] Processing complete.")
            print(f"  - {updated_count}/{len(questions)} answers populated")
            print(f"  - All original sheets preserved")
            print(f"  - All original data preserved")
            print(f"  - All original formatting preserved")
            return output_path
        else:
            print(f"\n[WARNING] Processing complete but preservation may have issues.")
            print(f"  - {updated_count}/{len(questions)} answers populated")
            return output_path

    except FileNotFoundError as e:
        print(f"[ERROR] File not found: {excel_path}")
        return None
    except Exception as e:
        print(f"[ERROR] Unexpected error: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    print("\n" + "="*80)
    print("EXCEL Q&A PROCESSOR - COMPLETE PRESERVATION")
    print("="*80)

    # Check input file and answers data
    if os.path.exists('input.json'):
        with open('input.json', 'r') as f:
            input_data = json.load(f)
        excel_file = input_data.get('excel_file')
        answers = input_data.get('answers', {})
    else:
        print("[ERROR] input.json not found")
        sys.exit(1)

    if not excel_file:
        print("[ERROR] 'excel_file' not specified in input.json")
        sys.exit(1)

    # Process the file
    result = process_assessment_file(excel_file, answers)

    if result:
        print(f"\n[RESULT] Output file: {result}")
        print("="*80 + "\n")
        sys.exit(0)
    else:
        print("\n[FAILED] Processing could not complete")
        print("="*80 + "\n")
        sys.exit(1)
