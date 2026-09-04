"""
build_assessment_snapshot.py

Builds a brand-new, standalone .xlsx workbook containing exactly one sheet,
named "Assessment", from a merged row set (existing Master Rubric rows plus
any newly answered/classified rows). This is never a patch of the source
file -- it is always a fresh workbook, and it never contains Cover, Change
Plan, Onboarding List, Product Backlog, Summary, or Renumber Map sheets.

Input (input.json in the working directory), shape:
{
  "rows": [
    {
      "dom_num": 1, "domain": "Platform Foundation", "cap_num": "1.1",
      "capability": "General System Access", "dimension": "Multi-system navigation",
      "priority": "P0", "discovery_question": "...", "branch_answer": "...",
      "townsq_capability": "...", "classification": "PC", "fb_priority": "",
      "assessor_notes": "...", "qid": "1.1.1", "uid": "Q0001"
    },
    ...
  ],
  "output_filename": "Assessment.xlsx"   # optional, defaults to Assessment.xlsx
}

Output: writes the .xlsx to the working directory (captured automatically
as a chat download) and prints a small JSON summary to stdout, including
the row count written -- the caller MUST verify this equals the expected
baseline row count before reporting success.
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "Dom #", "Domain", "Cap #", "Capability", "Dimension", "Priority",
    "Discovery Question", "Branch Answer", "TownSq Capability",
    "Classification", "FB Priority", "Assessor Notes", "QID", "UID",
]

FIELD_ORDER = [
    "dom_num", "domain", "cap_num", "capability", "dimension", "priority",
    "discovery_question", "branch_answer", "townsq_capability",
    "classification", "fb_priority", "assessor_notes", "qid", "uid",
]


def main():
    with open("input.json", "r", encoding="utf-8") as f:
        data = json.load(f)

    rows = data.get("rows", [])
    output_filename = data.get("output_filename", "Assessment.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment"

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_data in rows:
        ws.append([row_data.get(field, "") for field in FIELD_ORDER])

    # reasonable default column widths
    widths = [6, 20, 6, 22, 24, 8, 45, 45, 45, 14, 10, 45, 10, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    wb.save(output_filename)

    print(json.dumps({
        "output_filename": output_filename,
        "rows_written": len(rows),
        "sheet_name": ws.title,
    }))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)
